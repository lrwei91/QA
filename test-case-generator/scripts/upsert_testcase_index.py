#!/usr/bin/env python3
"""Create or update testcase and i18n indexes for validation artifacts."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from validate_i18n_json import analyze_i18n_json

VALID_EXTENSIONS = {
    '.md': 'md',
    '.xlsx': 'xlsx',
    '.docx': 'docx',
    '.csv': 'csv',
    '.txt': 'txt',
    '.json': 'json',
}
VALID_STATUSES = {'draft', 'active', 'archived'}
GENERIC_TAG_STOPWORDS = {'测试', '测试用例', '测试案例', '用例'}
GENERIC_SUFFIXES = ['汇总', '优化', '改造', '功能', '前后端']
CASES_DIR = 'generated'
I18N_DIR = 'i18n'
LEGACY_INDEX_NAME = 'index.json'
TESTCASE_INDEX_NAME = 'testcase-index.json'
I18N_INDEX_NAME = 'i18n-index.json'


def now_iso() -> str:
    return datetime.now().astimezone().replace(microsecond=0).isoformat()


def fail(message: str) -> int:
    print(f'ERROR: {message}')
    return 1


def stable_slug(text: str, prefix: str) -> str:
    cleaned = re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')
    digest = hashlib.sha1(text.encode('utf-8')).hexdigest()[:8]
    if cleaned:
        return f'{prefix}-{cleaned}-{digest}'
    return f'{prefix}-{digest}'


def default_index_data(index_type: str, index_path: Path) -> dict:
    if index_type == 'testcase':
        return {
            'version': 1,
            'store_name': 'QA Test Case Store',
            'root_dir': index_path.parent.name,
            'cases_dir': CASES_DIR,
            'updated_at': now_iso(),
            'entries': [],
        }
    return {
        'version': 1,
        'store_name': 'QA I18N JSON Store',
        'root_dir': index_path.parent.name,
        'i18n_dir': I18N_DIR,
        'updated_at': now_iso(),
        'entries': [],
    }


def normalize_index_data(index_data: dict, index_type: str, index_path: Path) -> dict:
    data = dict(index_data)
    defaults = default_index_data(index_type, index_path)
    for key, value in defaults.items():
        data.setdefault(key, value)
    return data


def load_index(index_path: Path, index_type: str) -> dict:
    if not index_path.exists():
        return default_index_data(index_type, index_path)
    return normalize_index_data(json.loads(index_path.read_text(encoding='utf-8')), index_type, index_path)


def dump_index(index_path: Path, data: dict) -> None:
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')


def first_non_empty_string(value) -> str:
    if isinstance(value, str):
        return value.strip()
    if value is None:
        return ''
    return str(value).strip()


def infer_xlsx_metadata(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    title = first_non_empty_string(worksheet['A1'].value) or path.stem

    platform_scope: list[str] = []
    seen_platforms: set[str] = set()
    max_row = min(worksheet.max_row, 5000)
    for row in range(8, max_row + 1):
        label = first_non_empty_string(worksheet.cell(row, 2).value)
        if label and label not in seen_platforms:
            seen_platforms.add(label)
            platform_scope.append(label)

    return {
        'title': title,
        'topic': title,
        'artifact_type': 'testcase',
        'platform_scope': platform_scope,
        'language_codes': [],
        'status': None,
    }


def infer_text_metadata(path: Path) -> dict:
    return {
        'title': path.stem,
        'topic': path.stem,
        'artifact_type': 'testcase',
        'platform_scope': [],
        'language_codes': [],
        'status': None,
    }


def infer_i18n_json_metadata(path: Path) -> dict:
    analysis = analyze_i18n_json(path)
    return {
        'title': analysis['name'],
        'topic': analysis['name'],
        'artifact_type': 'i18n-json',
        'platform_scope': [],
        'language_codes': analysis['language_codes'],
        'status': 'draft' if analysis['is_draft'] else None,
    }


def dedupe_keep_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        item = item.strip()
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def split_title_chunks(title: str) -> list[str]:
    return [chunk.strip() for chunk in re.split(r'[\-_/|｜]+', title) if chunk.strip()]


def strip_generic_suffixes(text: str) -> str:
    cleaned = text.strip().strip('()（）[]【】')
    previous = None
    while cleaned and cleaned != previous:
        previous = cleaned
        for suffix in GENERIC_SUFFIXES:
            if cleaned.endswith(suffix) and len(cleaned) > len(suffix):
                cleaned = cleaned[: -len(suffix)].rstrip().rstrip('-_/|｜').strip().strip('()（）[]【】')
                break
    return cleaned


def infer_module_from_name(name: str) -> str:
    for chunk in split_title_chunks(name):
        cleaned = strip_generic_suffixes(chunk)
        if cleaned and cleaned not in GENERIC_TAG_STOPWORDS and len(cleaned) >= 2:
            return cleaned
    cleaned = strip_generic_suffixes(name)
    return cleaned or name


def derive_tags(module: str, title: str, extra_tags: list[str]) -> list[str]:
    candidates = [module]
    for chunk in split_title_chunks(title):
        cleaned = strip_generic_suffixes(chunk)
        if not cleaned:
            continue
        if cleaned in GENERIC_TAG_STOPWORDS:
            continue
        if len(cleaned) < 2:
            continue
        candidates.append(cleaned)
    candidates.extend(extra_tags)
    return dedupe_keep_order(candidates)


def infer_module(rel_path: Path) -> str:
    parts = rel_path.parts
    if len(parts) >= 2 and parts[0] == CASES_DIR:
        return parts[1]
    if len(parts) >= 3 and parts[0] == I18N_DIR:
        return parts[1]
    if len(parts) >= 2 and parts[0] == I18N_DIR:
        return infer_module_from_name(rel_path.stem)
    if len(parts) >= 2:
        return parts[-2]
    return infer_module_from_name(rel_path.stem)


def build_group_key(module: str, topic: str) -> str:
    return stable_slug(f'{module}::{topic}', 'group')


def infer_artifact_metadata(path: Path) -> dict:
    file_format = VALID_EXTENSIONS[path.suffix.lower()]
    if file_format == 'xlsx':
        return infer_xlsx_metadata(path)
    if file_format == 'json':
        return infer_i18n_json_metadata(path)
    return infer_text_metadata(path)


def build_entry(
    path: Path,
    store_root: Path,
    args: argparse.Namespace,
) -> tuple[str, dict]:
    rel_path = path.relative_to(store_root)
    file_format = VALID_EXTENSIONS[path.suffix.lower()]
    inferred = infer_artifact_metadata(path)
    module = args.module or infer_module(rel_path)
    title = args.title or inferred['title']
    topic = args.topic or inferred['topic']
    module_ids = args.module_ids or [stable_slug(module, 'module')]
    status = args.status if args.status != 'auto' else (inferred.get('status') or 'active')
    tags = derive_tags(module, title, args.tags)
    base_entry = {
        'id': stable_slug(rel_path.as_posix(), 'tc'),
        'group_key': build_group_key(module, topic),
        'title': title,
        'module': module,
        'module_ids': dedupe_keep_order(module_ids),
        'topic': topic,
        'format': file_format,
        'rel_path': rel_path.as_posix(),
        'source_refs': dedupe_keep_order(args.source_refs),
        'tags': tags,
        'status': status,
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }

    if inferred['artifact_type'] == 'testcase':
        base_entry.update(
            {
                'platform_scope': dedupe_keep_order(args.platform_scope or inferred['platform_scope']),
                'template': args.template or '',
            }
        )
        return 'testcase', base_entry

    base_entry.update(
        {
            'language_codes': dedupe_keep_order(inferred['language_codes']),
            'template': args.template or '',
        }
    )
    return 'i18n-json', base_entry


def merge_entry(index_type: str, existing: dict | None, new_entry: dict) -> dict:
    if not existing:
        return new_entry

    merged = dict(new_entry)
    merged['id'] = existing.get('id', new_entry['id'])
    merged['created_at'] = existing.get('created_at') or new_entry['created_at']
    merged['source_refs'] = dedupe_keep_order(existing.get('source_refs', []) + new_entry['source_refs'])
    merged['tags'] = dedupe_keep_order(new_entry['tags'])
    merged['module_ids'] = dedupe_keep_order(new_entry['module_ids'])
    if index_type == 'testcase':
        merged['platform_scope'] = dedupe_keep_order(existing.get('platform_scope', []) + new_entry['platform_scope'])
    else:
        merged['language_codes'] = dedupe_keep_order(new_entry['language_codes'])
    return merged


def collect_files(store_root: Path, args: argparse.Namespace) -> list[Path]:
    if args.all:
        files = []
        for root_name in [CASES_DIR, I18N_DIR]:
            root_dir = store_root / root_name
            if not root_dir.exists():
                continue
            for path in sorted(root_dir.rglob('*')):
                if not path.is_file():
                    continue
                if path.name.startswith('.'):
                    continue
                if path.suffix.lower() in VALID_EXTENSIONS:
                    files.append(path.resolve())
        return files

    return [Path(item).resolve() for item in args.files]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Create or update testcase-index.json and i18n-index.json for validation artifacts.'
    )
    parser.add_argument('files', nargs='*', help='Artifact files to upsert into the indexes')
    parser.add_argument('--all', action='store_true', help='Scan the whole cases_dir and i18n_dir and upsert all supported files')
    parser.add_argument(
        '--testcase-index',
        default=f'testcases/{TESTCASE_INDEX_NAME}',
        help=f'Path to testcase index file. Defaults to testcases/{TESTCASE_INDEX_NAME}',
    )
    parser.add_argument(
        '--i18n-index',
        default=f'testcases/{I18N_INDEX_NAME}',
        help=f'Path to i18n index file. Defaults to testcases/{I18N_INDEX_NAME}',
    )
    parser.add_argument('--title', help='Override entry title')
    parser.add_argument('--topic', help='Override entry topic')
    parser.add_argument('--module', help='Override entry module')
    parser.add_argument('--module-id', action='append', dest='module_ids', default=[], help='Append module id')
    parser.add_argument('--platform', action='append', dest='platform_scope', default=[], help='Append platform label')
    parser.add_argument('--tag', action='append', dest='tags', default=[], help='Append tag')
    parser.add_argument('--source-ref', action='append', dest='source_refs', default=[], help='Append source ref')
    parser.add_argument('--template', default='', help='Set template name or path')
    parser.add_argument('--status', default='auto', help='Entry status: auto, draft, active, archived')
    return parser.parse_args()


def load_legacy_entries(store_root: Path) -> tuple[dict[str, dict], dict[str, dict]]:
    legacy_path = store_root / LEGACY_INDEX_NAME
    if not legacy_path.exists():
        return {}, {}

    try:
        data = json.loads(legacy_path.read_text(encoding='utf-8'))
    except json.JSONDecodeError:
        return {}, {}

    testcase_entries: dict[str, dict] = {}
    i18n_entries: dict[str, dict] = {}
    for entry in data.get('entries', []):
        rel_path = entry.get('rel_path')
        artifact_type = entry.get('artifact_type')
        if not isinstance(rel_path, str) or not rel_path:
            continue
        migrated = dict(entry)
        if artifact_type == 'i18n-json':
            migrated.pop('artifact_type', None)
            migrated.pop('platform_scope', None)
            i18n_entries[rel_path] = migrated
        else:
            migrated.pop('artifact_type', None)
            migrated.pop('language_codes', None)
            testcase_entries[rel_path] = migrated
    return testcase_entries, i18n_entries


def build_existing_maps(store_root: Path, testcase_index_path: Path, i18n_index_path: Path) -> tuple[dict[str, dict], dict[str, dict]]:
    legacy_testcase_entries, legacy_i18n_entries = load_legacy_entries(store_root)

    if testcase_index_path.exists():
        testcase_entries = {
            entry['rel_path']: entry
            for entry in load_index(testcase_index_path, 'testcase').get('entries', [])
        }
    else:
        testcase_entries = legacy_testcase_entries

    if i18n_index_path.exists():
        i18n_entries = {
            entry['rel_path']: entry
            for entry in load_index(i18n_index_path, 'i18n-json').get('entries', [])
        }
    else:
        i18n_entries = legacy_i18n_entries

    return testcase_entries, i18n_entries


def main() -> int:
    args = parse_args()
    if args.status not in VALID_STATUSES | {'auto'}:
        return fail(f"invalid status: {args.status}")

    if not args.all and not args.files:
        return fail('provide at least one artifact file or use --all')

    testcase_index_path = Path(args.testcase_index).resolve()
    i18n_index_path = Path(args.i18n_index).resolve()
    if testcase_index_path.parent != i18n_index_path.parent:
        return fail('testcase-index and i18n-index must be under the same testcases root')

    store_root = testcase_index_path.parent.resolve()
    files = collect_files(store_root, args)
    if not files:
        print('No artifact files found to index')
        return 0

    for path in files:
        if not path.exists():
            return fail(f'file not found: {path}')
        if path.suffix.lower() not in VALID_EXTENSIONS:
            return fail(f'unsupported file format: {path}')
        try:
            path.relative_to(store_root)
        except ValueError:
            return fail(f'file is outside testcase store root: {path}')
        if path.suffix.lower() == '.json':
            try:
                analyze_i18n_json(path)
            except ValueError as exc:
                return fail(f'invalid i18n json {path}: {exc}')

    testcase_index = load_index(testcase_index_path, 'testcase')
    i18n_index = load_index(i18n_index_path, 'i18n-json')
    existing_testcase_entries, existing_i18n_entries = build_existing_maps(store_root, testcase_index_path, i18n_index_path)
    updated_testcase_entries: dict[str, dict] = {}
    updated_i18n_entries: dict[str, dict] = {}

    for path in files:
        index_type, new_entry = build_entry(path, store_root, args)
        if index_type == 'testcase':
            existing = existing_testcase_entries.get(new_entry['rel_path'])
            updated_testcase_entries[new_entry['rel_path']] = merge_entry('testcase', existing, new_entry)
        else:
            existing = existing_i18n_entries.get(new_entry['rel_path'])
            updated_i18n_entries[new_entry['rel_path']] = merge_entry('i18n-json', existing, new_entry)

    for rel_path, entry in existing_testcase_entries.items():
        if rel_path in updated_testcase_entries:
            continue
        target = (store_root / rel_path).resolve()
        if target.exists():
            updated_testcase_entries[rel_path] = entry

    for rel_path, entry in existing_i18n_entries.items():
        if rel_path in updated_i18n_entries:
            continue
        target = (store_root / rel_path).resolve()
        if target.exists():
            updated_i18n_entries[rel_path] = entry

    testcase_index['entries'] = [updated_testcase_entries[key] for key in sorted(updated_testcase_entries)]
    testcase_index['updated_at'] = now_iso()
    dump_index(testcase_index_path, testcase_index)

    i18n_index['entries'] = [updated_i18n_entries[key] for key in sorted(updated_i18n_entries)]
    i18n_index['updated_at'] = now_iso()
    dump_index(i18n_index_path, i18n_index)

    print(
        f'Updated {len(files)} artifact file(s): '
        f'{len(updated_testcase_entries)} testcase entries -> {testcase_index_path}, '
        f'{len(updated_i18n_entries)} i18n entries -> {i18n_index_path}'
    )
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
