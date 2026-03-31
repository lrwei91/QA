#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
"""
generate_testcase_from_template.py - Generate test cases from template without cross-skill dependency.

Usage:
    python3 generate_testcase_from_template.py template.xlsx requirements.json output.xlsx

This script reduces dependency on Office-Skills/minimax-xlsx by providing
a standalone testcase generation utility.
"""

import argparse
import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment


REQUIRED_KEYS = {"平台", "模块", "功能点", "前置条件（测试点）", "操作步骤", "预期结果", "备注"}
TEMPLATE_FIXED_ROWS = 7  # Rows reserved for header info (adjust based on template)


def validate_inputs(template_xlsx: str, requirements_json: str, output_xlsx: str) -> tuple:
    """Validate input files exist and JSON structure is correct."""
    errors = []

    if not os.path.exists(template_xlsx):
        errors.append(f"Template xlsx file not found: {template_xlsx}")

    if not os.path.exists(requirements_json):
        errors.append(f"JSON file not found: {requirements_json}")

    output_dir = os.path.dirname(output_xlsx)
    if output_dir and not os.path.exists(output_dir):
        errors.append(f"Output directory not found: {output_dir}")

    if errors:
        return False, errors

    try:
        with open(requirements_json, 'r', encoding='utf-8') as f:
            requirements = json.load(f)
    except json.JSONDecodeError as e:
        return False, [f"Invalid JSON format: {e}"]
    except PermissionError:
        return False, [f"Permission denied reading JSON file: {requirements_json}"]
    except OSError as e:
        return False, [f"Error reading JSON file: {e}"]

    if not isinstance(requirements, dict):
        return False, ["JSON must contain an object with testcase data"]

    if "testcases" not in requirements or not isinstance(requirements["testcases"], list):
        return False, ["JSON must contain 'testcases' array"]

    for i, row in enumerate(requirements["testcases"]):
        if not isinstance(row, dict):
            return False, [f"Testcase {i+1} must be an object, got {type(row).__name__}"]
        missing = REQUIRED_KEYS - set(row.keys())
        if missing:
            return False, [f"Testcase {i+1} missing required keys: {', '.join(sorted(missing))}"]

    return True, requirements


def generate_from_template(template_xlsx: str, requirements_json: str, output_xlsx: str) -> int:
    """Generate test cases from template.

    Returns:
        0 on success, 1 on error
    """
    valid, result = validate_inputs(template_xlsx, requirements_json, output_xlsx)
    if not valid:
        for error in result:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1

    requirements = result
    testcases = requirements["testcases"]
    metadata = requirements.get("metadata", {})

    if not testcases:
        print("No test cases to generate")
        return 0

    try:
        wb = load_workbook(template_xlsx)
    except PermissionError:
        print(f"ERROR: Permission denied reading Excel file: {template_xlsx}", file=sys.stderr)
        return 1
    except OSError as e:
        print(f"ERROR: Failed to load Excel file: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"ERROR: Invalid Excel file format: {e}", file=sys.stderr)
        return 1

    ws = wb.active

    # Fill metadata if present
    metadata_mapping = {
        "测试平台": 2,  # B2
        "系统&版本": 4,  # D2
        "文档编写人": 2,  # B4
        "参考档": 4,  # D4
        "测试日期": 2,  # B6
        "最后更新": 4,  # D6
    }

    for meta_key, cell_col in metadata_mapping.items():
        if meta_key in metadata:
            cell_row = (metadata_mapping[meta_key] // 2) * 2  # Rows 2, 4, 6
            cell = ws.cell(row=cell_row, column=cell_col, value=str(metadata[meta_key]))

    # Column mapping
    column_map = {
        "序号": 1,
        "平台": 2,
        "模块": 3,
        "功能点": 4,
        "前置条件（测试点）": 5,
        "操作步骤": 6,
        "预期结果": 7,
        "测试结果": 8,
        "备注": 9,
    }

    wrap_columns = ["前置条件（测试点）", "操作步骤", "预期结果"]
    start_row = TEMPLATE_FIXED_ROWS + 1  # Data starts after fixed rows

    # Fill test cases
    for idx, testcase in enumerate(testcases, start=1):
        testcase["序号"] = str(idx)
        row_number = start_row + idx - 1

        for col_name, col_idx in column_map.items():
            value = testcase.get(col_name, "")
            cell = ws.cell(row=row_number, column=col_idx, value=value)

            if col_name in wrap_columns:
                cell.alignment = Alignment(wrap_text=True)

    # Save workbook
    try:
        wb.save(output_xlsx)
    except PermissionError:
        print(f"ERROR: Permission denied writing to Excel file: {output_xlsx}", file=sys.stderr)
        return 1
    except OSError as e:
        print(f"ERROR: Failed to save Excel file: {e}", file=sys.stderr)
        return 1

    print(f"Successfully generated {len(testcases)} test cases to {output_xlsx}")
    print(f"Data rows: {start_row} to {start_row + len(testcases) - 1}")
    return 0


def main():
    parser = argparse.ArgumentParser(
        description="Generate test cases from template without cross-skill dependency"
    )
    parser.add_argument("template_xlsx", help="Path to template xlsx file")
    parser.add_argument("requirements_json", help="Path to JSON file with test case data")
    parser.add_argument("output_xlsx", help="Path to output xlsx file")

    args = parser.parse_args()

    exit_code = generate_from_template(
        args.template_xlsx,
        args.requirements_json,
        args.output_xlsx
    )

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
