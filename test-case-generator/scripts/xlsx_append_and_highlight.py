#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
"""
xlsx_append_and_highlight.py — Append new testcase rows to an existing xlsx file with highlight.

Usage:
    python3 xlsx_append_and_highlight.py existing.xlsx new_rows.json output.xlsx --highlight yellow

Uses openpyxl for reliable xlsx handling.
"""

import argparse
import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment


def append_rows(existing_xlsx: str, new_rows_json: str, output_xlsx: str,
                start_row: int = None, highlight: bool = True, highlight_color: str = "FFFF00") -> None:
    """Append new rows to existing xlsx file using openpyxl."""

    # Load new rows from JSON
    with open(new_rows_json, 'r', encoding='utf-8') as f:
        new_rows = json.load(f)

    if not new_rows:
        print("No new rows to append")
        return

    # Load existing workbook
    wb = load_workbook(existing_xlsx)
    ws = wb.active

    # Find last data row (data starts at row 8)
    last_row = ws.max_row
    if start_row is None:
        current_row = last_row + 1
    else:
        current_row = start_row

    # Create yellow fill for highlighting
    yellow_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    # Column mapping for testcase data
    column_map = {
        "序号": 1,  # A
        "平台": 2,  # B
        "模块": 3,  # C
        "功能点": 4,  # D
        "前置条件（测试点）": 5,  # E
        "操作步骤": 6,  # F
        "预期结果": 7,  # G
        "测试结果": 8,  # H
        "备注": 9,  # I
    }

    # Wrap text columns
    wrap_columns = ["前置条件（测试点）", "操作步骤", "预期结果"]

    # Append new rows
    base_seq = last_row - 7  # Base sequence number (last existing sequence)
    row_number = current_row
    seq_number = base_seq + 1  # Start sequence from next number

    for row_data in new_rows:
        # Set sequential row number
        row_data["序号"] = str(seq_number)

        for col_name, col_idx in column_map.items():
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_number, column=col_idx, value=value)

            # Apply highlight fill to all cells in the row
            if highlight:
                cell.fill = PatternFill(
                    start_color=highlight_color,
                    end_color=highlight_color,
                    fill_type="solid"
                )

            # Apply text wrap for specific columns
            if col_name in wrap_columns:
                cell.alignment = Alignment(wrap_text=True)

        row_number += 1
        seq_number += 1

    # Save workbook
    wb.save(output_xlsx)

    print(f"Successfully appended {len(new_rows)} rows to {output_xlsx}")
    print(f"Data rows: 8 to {row_number - 1} (total: {row_number - 8} data rows)")


def main():
    parser = argparse.ArgumentParser(description="Append testcase rows to existing xlsx with highlight")
    parser.add_argument("existing_xlsx", help="Path to existing xlsx file")
    parser.add_argument("new_rows_json", help="Path to JSON file with new rows")
    parser.add_argument("output_xlsx", help="Path to output xlsx file")
    parser.add_argument("--start-row", type=int, help="Start row number (default: auto-detect)")
    parser.add_argument("--highlight", action="store_true", help="Highlight new rows (default: yellow)")
    parser.add_argument("--highlight-color", default="FFFF00", help="Highlight color (default: FFFF00)")

    args = parser.parse_args()

    append_rows(
        args.existing_xlsx,
        args.new_rows_json,
        args.output_xlsx,
        start_row=args.start_row,
        highlight=args.highlight,
        highlight_color=args.highlight_color
    )


if __name__ == "__main__":
    main()
