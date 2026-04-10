"""
Skill: Test Case Format - Python Implementation

Excel 导出与索引更新
"""

import json
import subprocess
import tempfile
import os
from pathlib import Path
from datetime import datetime
from typing import Any


def export_to_excel(
    test_cases: list[dict],
    output_dir: str = None,
    module_name: str = "默认模块",
    update_index: bool = True
) -> dict[str, Any]:
    """
    导出测试用例到 Excel

    参数:
        test_cases: 测试用例列表
        output_dir: 输出目录（可选）
        module_name: 模块名称
        update_index: 是否更新索引

    返回:
        {
            "success": bool,
            "excel_path": str | None,
            "index_updated": bool
        }
    """
    if not test_cases:
        return {"success": False, "error": "No test cases to export", "excel_path": None}

    # 确定输出路径
    if output_dir is None:
        output_dir = Path(__file__).parent.parent.parent / "testcases" / "generated"
    else:
        output_dir = Path(output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    # 生成文件名
    safe_module = _sanitize_filename(module_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = output_dir / f"{safe_module}_{timestamp}.xlsx"

    # 导出 Excel
    excel_path = _create_excel(test_cases, output_xlsx)

    if not excel_path:
        return {"success": False, "error": "Failed to create Excel", "excel_path": None}

    # 更新索引
    index_updated = False
    if update_index:
        index_updated = _update_index(excel_path, test_cases)

    return {
        "success": True,
        "excel_path": excel_path,
        "index_updated": index_updated
    }


def _sanitize_filename(filename: str) -> str:
    """清理文件名中的非法字符"""
    import re
    return re.sub(r'[^\w\u4e00-\u9fff\-]', '', filename)


def _create_excel(test_cases: list[dict], output_xlsx: Path) -> str | None:
    """创建 Excel 文件"""
    try:
        # 尝试使用模板脚本
        template_path = Path(__file__).parent.parent.parent / "templates" / "testcase_template.xlsx"

        if template_path.exists():
            rows_data = {"testcases": test_cases, "metadata": {}}

            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
                json.dump(rows_data, f, ensure_ascii=False, indent=2)
                temp_json = f.name

            try:
                script_path = Path(__file__).parent.parent / "scripts" / "generate_testcase_from_template.py"

                if script_path.exists():
                    result = subprocess.run(
                        ["python3", str(script_path), str(template_path), temp_json, str(output_xlsx)],
                        capture_output=True, text=True, timeout=60
                    )
                    if result.returncode == 0:
                        return str(output_xlsx)
            finally:
                try:
                    os.unlink(temp_json)
                except:
                    pass

        # 回退到 openpyxl
        return _create_excel_simple(test_cases, output_xlsx)

    except Exception:
        return _create_excel_simple(test_cases, output_xlsx)


def _create_excel_simple(test_cases: list[dict], output_xlsx: Path) -> str | None:
    """使用 openpyxl 简单创建 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        headers = ["序号", "平台", "模块", "功能点", "前置条件（测试点）", "操作步骤", "预期结果", "测试结果", "备注"]
        header_row = 8

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

        wrap_columns = ["前置条件（测试点）", "操作步骤", "预期结果"]

        for idx, testcase in enumerate(test_cases, 1):
            row_num = header_row + idx
            for col, header in enumerate(headers, 1):
                value = testcase.get(header, "")
                cell = ws.cell(row=row_num, column=col, value=value)
                if header in wrap_columns:
                    ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

        column_widths = [8, 12, 15, 20, 25, 25, 25, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(str(output_xlsx))
        return str(output_xlsx)

    except Exception:
        return None


def _update_index(excel_path: str, test_cases: list[dict]) -> bool:
    """更新测试用例索引"""
    try:
        script_path = Path(__file__).parent.parent / "scripts" / "upsert_testcase_index.py"

        if script_path.exists():
            result = subprocess.run(
                ["python3", str(script_path), excel_path],
                capture_output=True, text=True, timeout=60
            )
            return result.returncode == 0

        return False
    except Exception:
        return False
