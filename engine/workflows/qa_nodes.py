"""
QA Workflow Nodes - Lite Version (No API Key Required)

本工作流运行在 Claude Code 环境中，不需要配置 ANTHROPIC_API_KEY。
AI 相关的任务由当前对话的 Claude 直接完成，节点函数只负责数据处理。
"""

import json
from pathlib import Path
from typing import Any

from .qa_state import QAWorkflowState


def parse_input(state: QAWorkflowState) -> QAWorkflowState:
    """解析输入内容，检测是否包含 Axure/HTML 数据"""
    errors = state.get("errors", [])

    try:
        input_content = state.get("input_content", "")
        input_source = state.get("input_source", "text")

        if input_source == "file":
            file_path = Path(input_content)
            if not file_path.exists():
                errors.append(f"File not found: {input_content}")
                return {**state, "errors": errors}
            input_content = file_path.read_text(encoding="utf-8")

        elif input_source == "axure_dir":
            axure_dir = Path(input_content)
            if not axure_dir.is_dir():
                errors.append(f"Directory not found: {input_content}")
                return {**state, "errors": errors}
            axure_data = parse_axure_directory(axure_dir)
            return {**state, "axure_data": axure_data}

        return {**state, "input_content": input_content}

    except Exception as e:
        errors.append(f"parse_input failed: {str(e)}")
        return {**state, "errors": errors}


def parse_axure_directory(dir_path: Path) -> dict[str, Any]:
    """解析 Axure HTML 目录"""
    import subprocess

    script_path = Path(__file__).parent.parent / "scripts" / "parse_axure_html.py"
    result = subprocess.run(
        ["python3", str(script_path), str(dir_path), "--recursive", "--format", "testcase"],
        capture_output=True, text=True, encoding="utf-8"
    )

    if result.returncode != 0:
        return {"error": result.stderr, "modules": [], "function_points": []}

    try:
        data = json.loads(result.stdout)
        return data[0] if isinstance(data, list) else data
    except json.JSONDecodeError:
        return {"error": "Failed to parse JSON", "modules": [], "function_points": []}


def read_figma(state: QAWorkflowState) -> QAWorkflowState:
    """读取 Figma 设计稿 - 需要 FIGMA_ACCESS_TOKEN"""
    errors = state.get("errors", [])
    figma_url = state.get("figma_url")

    if not figma_url:
        return state

    try:
        import os, subprocess, re

        file_id_match = re.search(r'/file/([a-zA-Z0-9]+)/', figma_url)
        node_id_match = re.search(r'node-id=(\d+[:\-]\d+)', figma_url)

        if not file_id_match:
            errors.append("Invalid Figma URL")
            return {**state, "errors": errors, "figma_data": None}

        file_id = file_id_match.group(1)
        node_id = node_id_match.group(1) if node_id_match else ""

        figma_token = os.environ.get("FIGMA_ACCESS_TOKEN")
        if not figma_token:
            errors.append("FIGMA_ACCESS_TOKEN not set")
            return {**state, "errors": errors, "figma_data": None}

        api_endpoint = f"https://api.figma.com/v1/files/{file_id}/nodes?ids={node_id}" if node_id else f"https://api.figma.com/v1/files/{file_id}"

        result = subprocess.run(
            ["curl", "-s", "-H", f"X-Figma-Token: {figma_token}", api_endpoint],
            capture_output=True, text=True, timeout=30
        )

        if result.returncode != 0:
            errors.append(f"Figma API failed: {result.stderr}")
            return {**state, "errors": errors, "figma_data": None}

        api_response = json.loads(result.stdout)
        if "err" in api_response:
            errors.append(f"Figma API error: {api_response.get('status')}")
            return {**state, "errors": errors, "figma_data": None}

        figma_data = {
            "source": "figma_api",
            "file_id": file_id,
            "file_name": api_response.get("name", "Unknown"),
            "nodes": api_response.get("nodes", {})
        }

        return {**state, "figma_data": figma_data}

    except Exception as e:
        errors.append(f"read_figma failed: {str(e)}")
        return {**state, "errors": errors, "figma_data": None}


def generate_test_cases(state: QAWorkflowState) -> QAWorkflowState:
    """
    生成测试用例 - 准备阶段

    【重要】此节点不调用 AI，而是准备数据供 Claude Code 使用：
    1. 合并输入数据
    2. 构建 prompt
    3. 存储在 state 中，由 Claude Code 的 Claude 生成用例
    """
    errors = state.get("errors", [])

    try:
        input_content = state.get("input_content", "")
        axure_data = state.get("axure_data")
        figma_data = state.get("figma_data")
        case_type = state.get("case_type", "full")

        combined_input = {
            "requirement": input_content,
            "axure": axure_data,
            "figma": figma_data,
            "case_type": case_type
        }

        prompt = build_testcase_generation_prompt(combined_input)
        i18n_detected = detect_i18n(input_content, axure_data, figma_data)

        return {
            **state,
            "_prompt_for_claude": prompt,
            "_i18n_detected": i18n_detected
        }

    except Exception as e:
        errors.append(f"generate_test_cases failed: {str(e)}")
        return {**state, "errors": errors, "test_cases": [], "_i18n_detected": False}


def build_testcase_generation_prompt(combined_input: dict) -> str:
    """构建测试用例生成的 Prompt - 给当前对话的 Claude 使用"""
    requirement = combined_input.get("requirement", "")
    axure = combined_input.get("axure")
    figma = combined_input.get("figma")
    case_type = combined_input.get("case_type", "full")

    return f"""你是一个专业的测试工程师。请根据以下需求内容生成结构化测试用例。

## 需求内容
{requirement}

## Axure 设计数据
{json.dumps(axure, ensure_ascii=False) if axure else "无"}

## Figma 设计数据
{json.dumps(figma, ensure_ascii=False) if figma else "无"}

## 用例类型
{"冒烟用例" if case_type == "smoke" else "完整用例"}

## 输出格式要求
请以 JSON 数组格式输出测试用例列表，每条用例包含：
- 平台："客户端" 或 "账服"
- 模块：功能所属模块
- 功能点：具体验证目标
- 前置条件（测试点）：执行前必须满足的状态
- 操作步骤：可执行的操作序列
- 预期结果：明确可验证的结果
- 备注：用例类型标记

## 平台划分规则
- 客户端：页面展示、控件交互、文案提示、跳转、渲染、表单输入、按钮点击、弹窗、列表、禁用态、前端校验
- 账服：接口入参/出参、服务端校验、业务处理、状态变更、写库、查库、缓存、消息、异步任务、错误码、幂等、权限、风控、限流

请直接返回 JSON 数组。"""


def detect_i18n(input_content: str, axure_data: dict | None, figma_data: dict | None) -> bool:
    """检测是否包含多语言数据"""
    keywords = ["多语言", "国际化", "i18n", "en-us", "id-id", "pt-pt", "es-es", "bn-bn", "tr-tr", "fp-fp"]
    return any(kw.lower() in input_content.lower() for kw in keywords)


def check_i18n(state: QAWorkflowState) -> QAWorkflowState:
    """检查并生成多语言 JSON"""
    errors = state.get("errors", [])

    try:
        input_content = state.get("input_content", "")
        i18n_data = extract_i18n_data(input_content)

        if not i18n_data:
            return {**state, "i18n_json": None}

        standard_languages = ["en-us", "id-id", "pt-pt", "es-es", "bn-bn", "tr-tr", "fp-fp"]
        validation = validate_i18n_languages(i18n_data, standard_languages)

        if not validation["is_complete"]:
            errors.append(f"I18N incomplete: missing {validation['missing_languages']}")
            return {**state, "errors": errors, "i18n_json": None}

        i18n_json_path = save_i18n_json(i18n_data)
        return {**state, "i18n_json": i18n_data, "i18n_json_path": i18n_json_path}

    except Exception as e:
        errors.append(f"check_i18n failed: {str(e)}")
        return {**state, "errors": errors}


def extract_i18n_data(input_content: str) -> dict | None:
    """从输入内容中提取多语言数据"""
    import re
    standard_languages = ["en-us", "id-id", "pt-pt", "es-es", "bn-bn", "tr-tr", "fp-fp"]

    try:
        json_pattern = r'\{[^{}]*"?(?:en-us|id-id|pt-pt)"?[^{}]*\}'
        json_matches = re.findall(json_pattern, input_content, re.IGNORECASE)

        for match in json_matches:
            try:
                data = json.loads(match)
                if any(lang in data for lang in standard_languages):
                    return {"entries": [data], "source": "input_json"}
            except json.JSONDecodeError:
                continue
    except:
        pass

    return None


def validate_i18n_languages(i18n_data: dict, standard_languages: list[str]) -> dict:
    """校验多语言数据的语言完整性"""
    entries = i18n_data.get("entries", [])
    missing = []

    for entry in entries:
        for lang in standard_languages:
            if lang not in entry and lang not in missing:
                missing.append(lang)

    return {"is_complete": len(missing) == 0, "missing_languages": missing}


def save_i18n_json(i18n_data: dict) -> str | None:
    """保存多语言 JSON 文件"""
    from datetime import datetime

    try:
        output_dir = Path(__file__).parent.parent.parent / "testcases" / "i18n"
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"i18n_{timestamp}.json"

        output_data = {
            "created_at": datetime.now().isoformat(),
            "source": i18n_data.get("source", "unknown"),
            "entries": i18n_data.get("entries", [])
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

        return str(output_file)
    except Exception:
        return None


def export_excel(state: QAWorkflowState) -> QAWorkflowState:
    """导出 Excel 并更新索引"""
    errors = state.get("errors", [])

    try:
        test_cases = state.get("test_cases", [])
        export_option = state.get("export_option", "with_index")

        if not test_cases:
            return state

        module = test_cases[0].get("模块", "默认模块")
        import re
        safe_module = re.sub(r'[^\w\u4e00-\u9fff\-]', '', module)
        output_dir = Path(__file__).parent.parent.parent / "testcases" / "generated" / safe_module
        output_dir.mkdir(parents=True, exist_ok=True)

        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = output_dir / f"testcase_{timestamp}.xlsx"

        excel_path = create_excel_with_openpyxl(test_cases, output_xlsx)

        if not excel_path:
            errors.append("Failed to export Excel")
            return {**state, "errors": errors}

        return {**state, "excel_path": str(excel_path)}

    except Exception as e:
        errors.append(f"export_excel failed: {str(e)}")
        return {**state, "errors": errors}


def create_excel_with_openpyxl(test_cases: list[dict], output_xlsx: Path) -> str | None:
    """使用 openpyxl 创建 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        headers = ["序号", "平台", "模块", "功能点", "前置条件（测试点）", "操作步骤", "预期结果", "测试结果", "备注"]
        header_row = 8

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

        wrap_columns = ["前置条件（测试点）", "操作步骤", "预期结果"]

        for idx, testcase in enumerate(test_cases, 1):
            row_num = header_row + idx
            for col, header in enumerate(headers, 1):
                value = testcase.get(header, "")
                cell = ws.cell(row=row_num, column=col, value=value)
                if header in wrap_columns:
                    cell.alignment = Alignment(wrap_text=True)

        column_widths = [8, 12, 15, 20, 25, 25, 25, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(str(output_xlsx))
        return str(output_xlsx)
    except Exception:
        return None


# ==================== 条件判断函数 ====================

def should_read_figma(state: QAWorkflowState) -> bool:
    return state.get("figma_url") is not None

def has_i18n(state: QAWorkflowState) -> bool:
    return state.get("_i18n_detected", False)

def has_test_cases(state: QAWorkflowState) -> bool:
    return len(state.get("test_cases", [])) > 0


# ==================== 补充用例工作流节点 ====================

def load_existing_cases(state: QAWorkflowState) -> QAWorkflowState:
    """
    读取已有测试用例

    从 testcase-index.json 或指定文件路径读取已有用例
    """
    errors = state.get("errors", [])

    try:
        import json

        # 从 testcase-index.json 读取
        index_path = Path(__file__).parent.parent.parent / "testcases" / "testcase-index.json"

        if not index_path.exists():
            errors.append(f"Index file not found: {index_path}")
            return {**state, "errors": errors, "existing_cases": []}

        with open(index_path, 'r', encoding='utf-8') as f:
            index_data = json.load(f)

        entries = index_data.get("entries", [])

        # 如果有指定模块，过滤
        target_module = state.get("_target_module")
        if target_module:
            entries = [e for e in entries if e.get("module") == target_module]

        # 加载用例内容
        existing_cases = []
        for entry in entries:
            rel_path = entry.get("rel_path")
            if rel_path:
                case_path = Path(__file__).parent.parent.parent / rel_path
                if case_path.exists():
                    # 尝试读取 Excel 或 Markdown
                    if case_path.suffix == ".xlsx":
                        cases = read_testcase_excel(case_path)
                    elif case_path.suffix == ".md":
                        cases = read_testcase_markdown(case_path)
                    else:
                        cases = []
                    existing_cases.extend(cases)

        return {
            **state,
            "existing_cases": existing_cases,
            "existing_file_path": str(index_path)
        }

    except Exception as e:
        errors.append(f"load_existing_cases failed: {str(e)}")
        return {**state, "errors": errors, "existing_cases": []}


def read_testcase_excel(file_path: Path) -> list[dict]:
    """从 Excel 文件读取测试用例"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        cases = []
        headers = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 8:  # 表头在第 8 行
                headers = row
                continue
            if row_idx > 8 and headers:
                case = dict(zip(headers, row))
                if case.get("功能点"):  # 只读取有功能点的行
                    cases.append(case)

        wb.close()
        return cases

    except Exception:
        return []


def read_testcase_markdown(file_path: Path) -> list[dict]:
    """从 Markdown 文件读取测试用例"""
    try:
        content = file_path.read_text(encoding='utf-8')
        # 简单解析 Markdown 表格
        cases = []
        lines = content.split('\n')
        headers = None

        for line in lines:
            if line.startswith('|'):
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if headers is None:
                    headers = parts
                else:
                    case = dict(zip(headers, parts))
                    if case.get("功能点"):
                        cases.append(case)

        return cases

    except Exception:
        return []


def analyze_gaps(state: QAWorkflowState) -> QAWorkflowState:
    """
    分析已有用例的覆盖缺口

    基于 testcase-taxonomy.md 检查：
    - 正向主流程
    - 异常场景
    - 边界值
    - 状态流转
    - 权限角色
    """
    errors = state.get("errors", [])

    try:
        existing_cases = state.get("existing_cases", [])

        # 分析已有覆盖
        covered = {
            "positive": False,
            "exception": False,
            "boundary": False,
            "state_flow": False,
            "permission": False
        }

        for case in existing_cases:
            remark = case.get("备注", "").lower()
            if "功能" in remark or "p0" in remark:
                covered["positive"] = True
            if "异常" in remark or "exception" in remark:
                covered["exception"] = True
            if "边界" in remark or "boundary" in remark:
                covered["boundary"] = True
            if "状态" in remark or "flow" in remark:
                covered["state_flow"] = True
            if "权限" in remark or "permission" in remark:
                covered["permission"] = True

        # 识别缺口
        gaps = []
        if not covered["exception"]:
            gaps.append("缺少异常场景覆盖")
        if not covered["boundary"]:
            gaps.append("缺少边界值覆盖")
        if not covered["state_flow"]:
            gaps.append("缺少状态流转覆盖")
        if not covered["permission"]:
            gaps.append("缺少权限角色覆盖")

        return {
            **state,
            "identified_gaps": gaps,
            "_gap_analysis": covered
        }

    except Exception as e:
        errors.append(f"analyze_gaps failed: {str(e)}")
        return {**state, "errors": errors, "identified_gaps": []}


def generate_augment_cases(state: QAWorkflowState) -> QAWorkflowState:
    """
    生成补充用例

    基于缺口分析结果，生成补充用例
    此节点准备 prompt，由 Claude 生成实际用例
    """
    errors = state.get("errors", [])

    try:
        gaps = state.get("identified_gaps", [])
        existing_cases = state.get("existing_cases", [])
        input_content = state.get("input_content", "")

        # 构建补充用例的 prompt
        prompt = build_augment_prompt(existing_cases, gaps, input_content)

        return {
            **state,
            "_prompt_for_claude": prompt,
            "_augment_mode": True
        }

    except Exception as e:
        errors.append(f"generate_augment_cases failed: {str(e)}")
        return {**state, "errors": errors}


def build_augment_prompt(existing_cases: list[dict], gaps: list[str], input_content: str) -> str:
    """构建补充用例的 prompt"""
    return f"""你是一个专业的测试工程师。请基于以下已有用例和缺口分析，生成补充测试用例。

## 已有用例数量
{len(existing_cases)} 条

## 缺口分析
""" + "\n".join(f"- {gap}" for gap in gaps) + f"""

## 原始需求
{input_content if input_content else "见已有用例"}

## 输出要求
请生成补充用例，覆盖上述缺口。输出格式为 JSON 数组：
- 平台："客户端" 或 "账服"
- 模块：功能所属模块
- 功能点：具体验证目标
- 前置条件（测试点）：执行前必须满足的状态
- 操作步骤：可执行的操作序列
- 预期结果：明确可验证的结果
- 备注：用例类型标记

直接返回 JSON 数组。"""


def append_to_excel(state: QAWorkflowState) -> QAWorkflowState:
    """
    追加用例到 Excel 文件

    将新增用例追加到原文件末尾，新增行标黄
    """
    errors = state.get("errors", [])

    try:
        test_cases = state.get("test_cases", [])
        existing_file_path = state.get("existing_file_path")

        if not test_cases:
            return state

        if not existing_file_path:
            errors.append("No existing file path specified")
            return {**state, "errors": errors}

        # 调用追加脚本
        import subprocess
        import tempfile
        import json

        rows_data = {"testcases": test_cases, "metadata": {}}

        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
            json.dump(rows_data, f, ensure_ascii=False, indent=2)
            temp_json = f.name

        try:
            script_path = Path(__file__).parent.parent / "scripts" / "xlsx_append_and_highlight.py"

            if script_path.exists():
                result = subprocess.run(
                    ["python3", str(script_path), existing_file_path, temp_json, existing_file_path],
                    capture_output=True, text=True, timeout=60
                )
                if result.returncode != 0:
                    errors.append(f"Script failed: {result.stderr}")
            else:
                # 脚本不存在时使用 openpyxl 直接追加
                append_to_excel_simple(existing_file_path, test_cases)

        finally:
            try:
                import os
                os.unlink(temp_json)
            except:
                pass

        return state

    except Exception as e:
        errors.append(f"append_to_excel failed: {str(e)}")
        return {**state, "errors": errors}


def append_to_excel_simple(file_path: str, test_cases: list[dict]) -> bool:
    """简单追加用例到 Excel（无高亮）"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Alignment

        wb = load_workbook(file_path)
        ws = wb.active

        # 找到最后一行
        max_row = ws.max_row
        start_row = max_row + 1

        headers = ["序号", "平台", "模块", "功能点", "前置条件（测试点）", "操作步骤", "预期结果", "测试结果", "备注"]
        wrap_columns = ["前置条件（测试点）", "操作步骤", "预期结果"]

        for idx, testcase in enumerate(test_cases, 1):
            row_num = start_row + idx - 1
            for col, header in enumerate(headers, 1):
                value = testcase.get(header, "")
                ws.cell(row=row_num, column=col, value=value)
                if header in wrap_columns:
                    ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

        wb.save(file_path)
        wb.close()
        return True

    except Exception:
        return False


# ==================== 需求分析工作流节点 ====================

def analyze_requirements(state: QAWorkflowState) -> QAWorkflowState:
    """
    分析需求文档

    输出：
    - 模块拆分
    - 风险等级评估
    - 测试点清单
    - 关联影响模块
    """
    errors = state.get("errors", [])

    try:
        input_content = state.get("input_content", "")

        # 构建分析 prompt
        prompt = build_analysis_prompt(input_content)

        return {
            **state,
            "_prompt_for_claude": prompt,
            "_analysis_mode": True
        }

    except Exception as e:
        errors.append(f"analyze_requirements failed: {str(e)}")
        return {**state, "errors": errors}


def build_analysis_prompt(input_content: str) -> str:
    """构建需求分析 prompt"""
    return f"""你是一个专业的测试工程师。请分析以下需求文档，输出结构化分析结果。

## 需求内容
{input_content}

## 分析要求
请输出以下内容（JSON 格式）：

1. **模块拆分**: 识别需求涉及的主要功能模块
2. **风险等级**: 评估每个模块的风险等级 (P0/P1/P2)
3. **测试点清单**: 列出需要测试的功能点（不是完整用例）
4. **关联影响**: 识别可能影响的其他模块

## 输出格式
```json
{{
    "modules": [
        {{"name": "模块名", "risk": "P0/P1/P2", "test_points": ["测试点 1", "测试点 2"]}}
    ],
    "affected_modules": ["受影响模块 1", "受影响模块 2"],
    "summary": "分析总结"
}}
```

直接返回 JSON。"""


def output_analysis(state: QAWorkflowState) -> QAWorkflowState:
    """
    输出需求分析结果

    将分析结果格式化输出
    """
    errors = state.get("errors", [])

    try:
        analysis_result = state.get("_analysis_result", {})

        # 这里只是标记分析完成
        # 实际输出由 Claude Code 格式化展示

        return {
            **state,
            "_analysis_completed": True
        }

    except Exception as e:
        errors.append(f"output_analysis failed: {str(e)}")
        return {**state, "errors": errors}


# ==================== 并行执行版本 ====================

async def parse_input_parallel(state: QAWorkflowState) -> QAWorkflowState:
    """
    并行解析输入内容和读取 Figma

    同时执行：
    1. parse_input (解析文本/Axure)
    2. read_figma (读取设计稿)
    """
    from .utils import run_async, ParallelExecutor

    executor = ParallelExecutor()

    # 添加并行任务
    executor.add_task(parse_input, state)

    # 如果有 Figma URL，添加读取任务
    if state.get("figma_url"):
        executor.add_task(read_figma, state)

    return executor.run(state)


# ==================== 带重试的节点版本 ====================

from .utils import with_retry

@with_retry(max_retries=3, delay=1.0)
def parse_input_with_retry(state: QAWorkflowState) -> QAWorkflowState:
    """parse_input with retry mechanism"""
    return parse_input(state)


@with_retry(max_retries=3, delay=2.0)
def read_figma_with_retry(state: QAWorkflowState) -> QAWorkflowState:
    """read_figma with retry mechanism (longer delay for API calls)"""
    return read_figma(state)


@with_retry(max_retries=2, delay=1.0)
def export_excel_with_retry(state: QAWorkflowState) -> QAWorkflowState:
    """export_excel with retry mechanism"""
    return export_excel(state)
