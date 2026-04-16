#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
"""
snapshot_testcase.py — Extract testcase data from xlsx and save as snapshot.

Usage:
    python3 snapshot_testcase.py outputs/generated/运营活动/复充返利活动.xlsx
    python3 snapshot_testcase.py outputs/generated/运营活动/复充返利活动.xlsx --output-dir outputs/snapshots/复充返利活动/

This script reads testcase data from an Excel file and saves it as a JSON snapshot.
The snapshot can be used for:
- Theme page test point summary generation
- Version tracking and diff comparison
- Knowledge base query enrichment
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Standard column names (possible variations)
COLUMN_NAME_VARIANTS = {
    'id': ['序号', 'ID', '编号'],
    'platform': ['平台', '测试平台', '端'],
    'module': ['模块', '功能模块', '系统'],
    'function': ['功能点', '功能', '测试功能'],
    'precondition': ['前置条件', '前置条件（测试点）', '测试点', '前提条件'],
    'steps': ['操作步骤', '步骤', '操作', '测试步骤'],
    'expected': ['预期结果', '预期', '期望结果', '预计结果'],
    'result': ['测试结果', '结果', '实际结果'],
    'remark': ['备注', '注释', '说明'],
}

# Standard column order (0-indexed)
STANDARD_COLUMNS = ['id', 'platform', 'module', 'function', 'precondition', 'steps', 'expected', 'result', 'remark']

METADATA_ROW = 2  # Metadata starts from row 2


def detect_header_row(ws, max_search_row: int = 15) -> tuple[int, dict[str, int]]:
    """Detect the header row and column mapping.

    Args:
        ws: Worksheet object
        max_search_row: Maximum row to search for header

    Returns:
        tuple: (header_row_index, column_map)
        - header_row_index: 1-indexed row number where header is found
        - column_map: dict mapping field names to column indices (1-indexed)
    """
    # Search for the header row by looking for "序号" or "平台"
    header_row = None
    for row in range(1, min(max_search_row + 1, ws.max_row + 1)):
        row_values = [str(ws.cell(row, col).value or '') for col in range(1, 15)]
        # Check if this row contains header keywords
        if any(h in row_values for h in ['序号', '平台']):
            header_row = row
            # Build column mapping
            column_map = {}
            for col_idx, value in enumerate(row_values, start=1):
                value = value.strip()
                for field_name, variants in COLUMN_NAME_VARIANTS.items():
                    if value in variants and field_name not in column_map:
                        column_map[field_name] = col_idx
            # Fill in missing columns with defaults based on standard order
            if 'id' in column_map:
                base_col = column_map['id']
                for i, field in enumerate(STANDARD_COLUMNS):
                    if field not in column_map:
                        expected_col = base_col + i
                        if expected_col <= len(row_values):
                            column_map[field] = expected_col
            return header_row, column_map

    # Fallback: assume standard format with header at row 7 or 8
    for row in [7, 8]:
        if row <= ws.max_row:
            first_cell = str(ws.cell(row, 1).value or '')
            if '序号' in first_cell:
                column_map = {field: idx + 1 for idx, field in enumerate(STANDARD_COLUMNS)}
                return row, column_map

    # Last resort: use default mapping
    return 8, {field: idx + 1 for idx, field in enumerate(STANDARD_COLUMNS)}


def extract_metadata(ws, max_row: int = 7) -> dict:
    """Extract metadata from the top section of the worksheet.

    Metadata cells:
    - B2: 测试平台
    - F2: 测试日期
    - B3: 系统&版本
    - F3: 最后更新
    - B4: 文档编写人
    - F4: 文档测试人
    - B5: 策划负责人
    - F5: 审阅人员
    - B6: 参考档
    """
    metadata_map = {
        '测试平台': (2, 2),  # B2
        '测试日期': (2, 6),  # F2
        '系统&版本': (3, 2),  # B3
        '最后更新': (3, 6),  # F3
        '文档编写人': (4, 2),  # B4
        '文档测试人': (4, 6),  # F4
        '策划负责人': (5, 2),  # B5
        '审阅人员': (5, 6),  # F5
        '参考档': (6, 2),    # B6
    }

    metadata = {}
    for key, (row, col) in metadata_map.items():
        if row <= max_row:
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value:
                metadata[key] = str(value).strip()

    return metadata


def extract_testcases_from_xlsx(xlsx_path: Path) -> tuple[list[dict], dict]:
    """读取 Excel 文件，提取测试用例数据和元数据。

    Returns:
        tuple: (testcases list, metadata dict)
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    ws = wb.active

    # Detect header row and column mapping
    header_row, column_map = detect_header_row(ws)

    # Extract metadata from top section
    metadata = extract_metadata(ws, max_row=header_row)

    # Extract testcase data starting from row after header
    testcases = []
    for row in range(header_row + 1, ws.max_row + 1):
        # Skip empty rows
        first_cell = ws.cell(row, column_map.get('id', 1))
        first_value = first_cell.value
        if not first_value:
            continue
        # Skip non-numeric ID values (like "序号" header itself)
        try:
            if isinstance(first_value, str) and not first_value.strip().isdigit():
                continue
        except Exception:
            continue

        testcase = {
            'id': int(float(first_value)) if first_value else len(testcases) + 1,
            'platform': str(ws.cell(row, column_map.get('platform', 2)).value or '').strip(),
            'module': str(ws.cell(row, column_map.get('module', 3)).value or '').strip(),
            'function': str(ws.cell(row, column_map.get('function', 4)).value or '').strip(),
            'precondition': str(ws.cell(row, column_map.get('precondition', 5)).value or '').strip(),
            'steps': str(ws.cell(row, column_map.get('steps', 6)).value or '').strip(),
            'expected': str(ws.cell(row, column_map.get('expected', 7)).value or '').strip(),
            'result': str(ws.cell(row, column_map.get('result', 8)).value or '').strip(),
            'remark': str(ws.cell(row, column_map.get('remark', 9)).value or '').strip(),
        }
        testcases.append(testcase)

    wb.close()
    return testcases, metadata


def build_test_point_summary(testcases: list[dict]) -> dict:
    """从测试用例生成测试点摘要。

    Groups test points by platform (server/client) and function.
    """
    summary = {
        'server_side': [],
        'client_side': [],
        'all_points': []
    }

    seen_points = set()

    for tc in testcases:
        # Use function point or module as the test point
        point = tc['function'] or tc['module'] or f"用例 #{tc['id']}"

        # Skip duplicates
        point_key = f"{tc['platform']}:{point}"
        if point_key in seen_points:
            continue
        seen_points.add(point_key)

        # Build test point entry
        point_entry = {
            'id': tc['id'],
            'point': point,
            'module': tc['module'],
        }

        if tc['platform'] == '账服':
            summary['server_side'].append(point)
        elif tc['platform'] == '客户端':
            summary['client_side'].append(point)

        summary['all_points'].append(point_entry)

    return summary


def save_snapshot(xlsx_path: Path, output_dir: Path, dry_run: bool = False) -> Path:
    """保存快照文件。

    Args:
        xlsx_path: Path to the Excel file
        output_dir: Directory to save the snapshot
        dry_run: If True, only print info without writing

    Returns:
        Path to the saved snapshot file
    """
    # Extract data
    testcases, metadata = extract_testcases_from_xlsx(xlsx_path)

    # Build summary
    test_point_summary = build_test_point_summary(testcases)

    # Build snapshot structure
    timestamp = datetime.now().astimezone().isoformat()
    snapshot = {
        'version': 1,
        'snapshot_at': timestamp,
        'source_file': str(xlsx_path),
        'metadata': {
            'title': xlsx_path.stem,
            'module': xlsx_path.parent.name,
            'platform_scope': list(set(tc['platform'] for tc in testcases if tc['platform'])),
            'total_cases': len(testcases),
            'extra_metadata': metadata,
        },
        'testcases': testcases,
        'test_point_summary': test_point_summary,
        'statistics': {
            'server_side_count': len(test_point_summary['server_side']),
            'client_side_count': len(test_point_summary['client_side']),
        }
    }

    if dry_run:
        ts = timestamp.replace(':', '-')
        print(f"Would save snapshot to: {output_dir / f'{ts}.json'}")
        print(f"Total testcases: {len(testcases)}")
        print(f"Server side: {snapshot['statistics']['server_side_count']}")
        print(f"Client side: {snapshot['statistics']['client_side_count']}")
        return None

    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    # Save snapshot with timestamp filename
    snapshot_file = output_dir / f"{timestamp.replace(':', '-')}.json"
    snapshot_file.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding='utf-8')

    return snapshot_file


def get_latest_snapshot(snapshot_dir: Path) -> Optional[Path]:
    """Get the latest snapshot file from the snapshot directory.

    Returns:
        Path to the latest snapshot file, or None if no snapshots exist
    """
    if not snapshot_dir.exists():
        return None

    snapshot_files = sorted(snapshot_dir.glob('*.json'), reverse=True)
    return snapshot_files[0] if snapshot_files else None


def main() -> int:
    parser = argparse.ArgumentParser(description='Extract testcase data from xlsx and save as snapshot')
    parser.add_argument('xlsx_file', type=Path, help='Path to the Excel file')
    parser.add_argument('--output-dir', '-o', type=Path,
                        default=Path('outputs/snapshots'),
                        help='Directory to save snapshots (default: outputs/snapshots/<module>/)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Show what would be done without actually writing')
    parser.add_argument('--show', '-s', action='store_true',
                        help='Show snapshot content instead of saving')

    args = parser.parse_args()

    # Validate input file
    if not args.xlsx_file.exists():
        print(f"ERROR: File not found: {args.xlsx_file}", file=sys.stderr)
        return 1

    if not args.xlsx_file.suffix.lower() == '.xlsx':
        print(f"ERROR: Not an Excel file: {args.xlsx_file}", file=sys.stderr)
        return 1

    # Determine output directory
    if args.output_dir == Path('outputs/snapshots'):
        # Default: create subdirectory based on module and filename
        module = args.xlsx_file.parent.name
        snapshot_dir = args.output_dir / module / args.xlsx_file.stem
    else:
        snapshot_dir = args.output_dir

    # Show mode: display what would be extracted
    if args.show:
        testcases, metadata = extract_testcases_from_xlsx(args.xlsx_file)
        summary = build_test_point_summary(testcases)

        print(f"=== {args.xlsx_file.name} ===")
        print(f"Module: {args.xlsx_file.parent.name}")
        print(f"Total testcases: {len(testcases)}")
        print(f"Server side: {len(summary['server_side'])}")
        print(f"Client side: {len(summary['client_side'])}")
        print()
        print("Server side test points:")
        for point in summary['server_side'][:10]:
            print(f"  - {point}")
        if len(summary['server_side']) > 10:
            print(f"  ... and {len(summary['server_side']) - 10} more")
        print()
        print("Client side test points:")
        for point in summary['client_side'][:10]:
            print(f"  - {point}")
        if len(summary['client_side']) > 10:
            print(f"  ... and {len(summary['client_side']) - 10} more")
        return 0

    # Save snapshot
    snapshot_file = save_snapshot(args.xlsx_file, snapshot_dir, dry_run=args.dry_run)

    if snapshot_file:
        print(f"Snapshot saved: {snapshot_file}")
        print(f"Total testcases: {len(extract_testcases_from_xlsx(args.xlsx_file)[0])}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
